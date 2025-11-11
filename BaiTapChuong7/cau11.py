
import openpyxl
from openpyxl.utils import get_column_letter
import os

# Tên file Excel để lưu trữ dữ liệu
FILE_NAME = 'employee_data.xlsx'
SHEET_NAME = 'DanhSachNhanVien'

# Cấu trúc dữ liệu trong bộ nhớ (List chứa Dictionaries)
# employees = [ {'Ma': 'NV001', 'Ten': 'An', 'Tuoi': 18}, ... ]
employees = []

# --- HÀM XỬ LÝ FILE EXCEL (I/O) ---

def load_data():
    """Đọc danh sách Nhân viên từ File Excel."""
    global employees
    employees = []
    
    if not os.path.exists(FILE_NAME):
        print(f"File '{FILE_NAME}' không tồn tại. Bắt đầu với danh sách trống.")
        return

    try:
        workbook = openpyxl.load_workbook(FILE_NAME)
        sheet = workbook[SHEET_NAME]
        
        # Bắt đầu đọc từ Hàng 2 (bỏ qua hàng tiêu đề)
        for row_index in range(2, sheet.max_row + 1):
            
            ma = sheet.cell(row=row_index, column=2).value # Cột B
            ten = sheet.cell(row=row_index, column=3).value # Cột C
            tuoi = sheet.cell(row=row_index, column=4).value # Cột D
            
            # Kiểm tra dữ liệu hợp lệ (không phải dòng trống)
            if ma is not None and ten is not None and tuoi is not None:
                try:
                    employees.append({
                        'Ma': str(ma),
                        'Ten': str(ten),
                        'Tuoi': int(tuoi)
                    })
                except ValueError:
                    # Bỏ qua dòng nếu Tuổi không phải số
                    continue 

        print(f" Đã tải thành công {len(employees)} nhân viên từ file.")
        
    except KeyError:
        print(f"Không tìm thấy sheet có tên '{SHEET_NAME}'. Bắt đầu với danh sách trống.")
    except Exception as e:
        print(f"Lỗi khi đọc file Excel: {e}. Bắt đầu với danh sách trống.")

def save_data():
    """Lưu danh sách Nhân viên vào File Excel."""
    
    # 1. Khởi tạo Workbook và Sheet
    if os.path.exists(FILE_NAME):
        workbook = openpyxl.load_workbook(FILE_NAME)
        # Xóa sheet cũ nếu tồn tại
        if SHEET_NAME in workbook.sheetnames:
            workbook.remove(workbook[SHEET_NAME])
    else:
        workbook = openpyxl.Workbook()
        # Xóa sheet mặc định nếu nó trống
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])
            
    sheet = workbook.create_sheet(SHEET_NAME, 0) # Tạo sheet mới ở vị trí đầu tiên

    # 2. Ghi tiêu đề (Hàng 1)
    headers = ['STT', 'Mã', 'Tên', 'Tuổi']
    sheet.append(headers)

    # 3. Ghi dữ liệu (Từ Hàng 2)
    for index, emp in enumerate(employees):
        sheet.append([
            index + 1, # STT
            emp['Ma'],
            emp['Ten'],
            emp['Tuoi']
        ])

    # 4. Tự động điều chỉnh độ rộng cột cho đẹp
    for col in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 15

    # 5. Lưu Workbook
    try:
        workbook.save(FILE_NAME)
        print(f"Đã lưu {len(employees)} nhân viên vào file '{FILE_NAME}' thành công.")
    except Exception as e:
        print(f"Lỗi khi lưu file: {e}")

# --- CÁC HÀM XỬ LÝ NGHIỆP VỤ ---

def display_employee_list():
    """Hiển thị danh sách nhân viên trong bộ nhớ."""
    if not employees:
        print("Danh sách nhân viên trống.")
        return

    print("\n--- DANH SÁCH NHÂN VIÊN ---")
    print(f"{'STT':<5} {'Mã':<8} {'Tên':<20} {'Tuổi':<5}")
    print("-" * 40)
    for index, emp in enumerate(employees):
        print(f"{index + 1:<5} {emp['Ma']:<8} {emp['Ten']:<20} {emp['Tuoi']:<5}")
    print("-" * 40)

def add_employee():
    """Thêm mới một nhân viên vào danh sách."""
    print("\n--- THÊM MỚI NHÂN VIÊN ---")
    
    ma = input("Nhập Mã nhân viên: ").strip()
    ten = input("Nhập Tên nhân viên: ").strip()
    
    while True:
        try:
            tuoi = int(input("Nhập Tuổi: "))
            if 0 < tuoi < 150:
                break
            else:
                print("Tuổi không hợp lệ.")
        except ValueError:
            print("Vui lòng nhập tuổi là một số nguyên.")

    # Kiểm tra trùng Mã (tùy chọn)
    if any(emp['Ma'] == ma for emp in employees):
        print(" Mã nhân viên đã tồn tại. Thao tác thêm bị hủy.")
        return
        
    employees.append({
        'Ma': ma,
        'Ten': ten,
        'Tuoi': tuoi
    })
    print(f"Nhân viên '{ten}' ({tuoi} tuổi) đã được thêm vào bộ nhớ.")

def sort_employees_by_age():
    """Sắp xếp danh sách Nhân viên theo Tuổi tăng dần."""
    
    if not employees:
        print("Danh sách trống, không thể sắp xếp.")
        return
        
    # Sắp xếp theo 'Tuoi' tăng dần (reverse=False là mặc định)
    employees.sort(key=lambda emp: emp['Tuoi'])
    
    print("\n--- SẮP XẾP THÀNH CÔNG ---")
    print("Danh sách đã được sắp xếp theo Tuổi TĂNG DẦN.")
    display_employee_list()

# --- CHƯƠNG TRÌNH CHÍNH (MENU) ---

def main_menu():
    """Hiển thị menu chính và xử lý lựa chọn."""
    
    load_data() # Tải dữ liệu khi chương trình khởi động
    
    while True:
        print("\n==================================")
        print("    PHẦN MỀM QUẢN LÝ NHÂN VIÊN")
        print("    Lưu trữ bằng File Excel")
        print("==================================")
        print("1. Thêm mới Nhân viên (vào bộ nhớ)")
        print("2. Xem danh sách Nhân viên")
        print("3. Sắp xếp Nhân viên theo Tuổi (Tăng dần)")
        print("4. Lưu dữ liệu vào Excel và Thoát")
        
        choice = input("Nhập lựa chọn của bạn (1-4): ").strip()
        
        if choice == '1':
            add_employee()
        elif choice == '2':
            display_employee_list()
        elif choice == '3':
            sort_employees_by_age()
        elif choice == '4':
            save_data()
            print("Cảm ơn bạn đã sử dụng chương trình!")
            break
        else:
            print("Lựa chọn không hợp lệ. Vui lòng chọn từ 1 đến 4.")

# Khởi chạy chương trình (Vui lòng bỏ comment dòng dưới để chạy)
main_menu()
